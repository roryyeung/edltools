# -*- coding: utf-8 -*-
import timecode
import re
from openpyxl import Workbook
from openpyxl.styles import Font
import os

def importEdlTitle(filepath):
    """Helper function - extracts the title from an File129 formatted EDL and returns a string"""
    with open(filepath) as file:
        for line in file.readlines():
            if line.startswith("TITLE") == True:
                return line[6:].strip()

def importEdlFcm(filepath):
    """Helper function - extracts the title from an File129 formatted EDL and returns a string"""
    with open(filepath) as file:
        for line in file.readlines():
            if line.startswith("FCM") == True:
                return line[5:].strip()

def importEdlBody(filepath,frameRate=25):
    """Helper function - extracts the body from an File129 formatted EDL and returns a list of dicts"""
    body = []
    with open(filepath) as file:
        counter = -1
        for line in file.readlines():
            
            # This section of logic eliminates non-body lines
            if line.strip() == "":
                continue
            if line.startswith("TITLE") == True:
                continue
            if line.startswith("FCM") == True:
                continue
            if line.startswith("* =") == True:
                #This logic detects the end of section marker and breaks the read loop
                return body
            # This section changes between clip and effect lines on the EDL
            if line.startswith("*") == False:
                # This section addes clip lines to the EDL body
                counter = counter + 1
                clip = line.split()
                # This section converts timecode to timecode objects
                i = 0
                while i < len(clip):
                    if re.search("[0-9]:[0-9]:[0-9]:[0-9]",clip[i]):
                        clip[i] = Timecode(frameRate,clip[i])
                    i +=1
                body.append({"clip":clip})
            else:
                # This section adds effect lines to the EDL body
                line = line.strip("*")
                if ":" in line:
                    key, value = line.split(":")
                    value = value.strip()
                    body[counter][key] = value
                else:
                    continue
    return body

def dumpEffects():
    """Helper function - removes all effects (non-clip) lines from the EDL."""
    # Todo
    pass

def exportXls(body,path):
    """Helper fucntcion - accepts a body object, a output path and an output file name."""
    # See tutorial at https://realpython.com/openpyxl-excel-spreadsheets-python/
    workbook = Workbook()
    sheet = workbook.active

    # Get the Headers
    headers = list(body[0].keys())

    # Set Headers in Excel
    i=0
    while i < len(headers):
        celref = sheet.cell(row=1,column=i+1)
        celref.value = headers[i]
        celref.font = Font(bold=True)
        i += 1

    # Create body in Excel
    i=0
    while i < len(body):
        j=0
        while j < len(headers):
            content = str(body[i].get(headers[j]))
            celref = sheet.cell(row=i+2,column = j+1)
            celref.value = content
            j += 1
        i += 1
    
    # # Test this a lot!


    # # Create File Path using OS
    # joinedPath = os.path.join(path,fileName)
    # print(joinedPath)

    # Save workbook
    workbook.save(filename=path)